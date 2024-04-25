import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def allocateTasksToWeeks(tasks):
    colWeekHours = [40] #first column, representing a work week, or 40 hours
    chronogram = [] 

    for task in tasks:
        weeks = len(colWeekHours)
        taskRow = ['_'] * weeks #current row times the weeks needed 
        while task > 0: #While the task has hours left to assing
            for i in range(len(colWeekHours)):
                if task <= colWeekHours[i]: #task needs less hours than available in current work week
                    colWeekHours[i] -= task
                    taskRow[i] = 'X'
                    task = 0 #Task hours fully allocated
                    break
                else:        #Task needs more hours than available in current work week
                    if colWeekHours[i] > 0:
                        task -= colWeekHours[i]
                        taskRow[i] = 'X'
                        colWeekHours[i] = 0

            #If task still has hours left not allocated, add new week
            if task > 0:   
                colWeekHours.append(40)
                taskRow.append('_')    #Extend task row for the new week

        #Update chronogram with next task
        chronogram.append(taskRow) 

         #Add weeks not used by tasks 
        for row in chronogram: 
            while len(row) < len(taskRow):
                row.append('_')

    return chronogram

def chronogramToExcel(chronogram, filename="chronogram.xlsx"):
    #Create DataFrame from chronogram
    df = pd.DataFrame(chronogram)

    #Write DataFrame to Excel file
    df.to_excel(filename, index=False, header=False)

    #Open Excel file and color cells
    wb = Workbook()
    ws = wb.active #Get current active sheet
    for index, row in enumerate(chronogram, start=1):
        for col_index, value in enumerate(row, start=1):
            if value == 'X':  #If cell has X
                cell = ws.cell(row=index, column=col_index)

                #Replace X, and fill cell with solid orange color
                cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    wb.save(filename) #Save Excel file updated

    df.to_csv("chronogram.csv", index=False)#Convert Excel file to csv

#Ask user for input (hours as separated values by comma)
taskHoursInput = input("Add tasks hours (as comma-separated values): ")

#Convert input string removing extra spaces and split input into a list of integers
tasks = [int(x.strip()) for x in taskHoursInput.split(',')]

#Generate the chronogram from user input
chronogram = allocateTasksToWeeks(tasks)

#Save chronogram to Excel with colored cells format
chronogramToExcel(chronogram, "chronogram.xlsx")




        


