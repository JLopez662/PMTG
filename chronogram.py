import pandas as pd
import os
import re
import time
from datetime import timedelta, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule
from copy import copy

def is_file_open(file_path):
    try:
        # Try to open the file in append mode
        with open(file_path, 'a'):
            pass
    except IOError:
        # If an IOError is raised, it means the file is open
        return True
    return False

def format_blank_cells(ws, rows=100, cols=50):
    for row in range(1, rows + 1):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            cell.border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))

def create_sheet_copy(wb, source_sheet_name, target_sheet_name):
    source_sheet = wb[source_sheet_name]
    target_sheet = wb.copy_worksheet(source_sheet)
    target_sheet.title = target_sheet_name
    return target_sheet

def allocateTasksToWeeks(milestones_tasks):
    chronogram = []
    last_end_week = 0

    for milestone_name, tasks in milestones_tasks:
        if len(chronogram) > 0:
            last_end_week += 1

        colWeekHours = [40.0] * (last_end_week + 1)
        milestone_rows = []

        for task in tasks:
            initial_task_hours = task
            weeks = len(colWeekHours)
            taskRow = ['_'] * weeks
            while task > 0:
                for i in range(last_end_week, len(colWeekHours)):
                    if task <= colWeekHours[i]:
                        colWeekHours[i] -= task
                        taskRow[i] = 'X'
                        task = 0
                        last_end_week = max(last_end_week, i)
                        break
                    else:
                        if colWeekHours[i] > 0:
                            task -= colWeekHours[i]
                            taskRow[i] = 'X'
                            colWeekHours[i] = 0
                            last_end_week = max(last_end_week, i)

                if task > 0:
                    colWeekHours.append(40.0)
                    taskRow.append('_')

            milestone_rows.append(taskRow)

        chronogram.extend(milestone_rows)
        if milestone_name != milestones_tasks[-1][0]:
            chronogram.append([''] * len(colWeekHours))

    return chronogram

all_week_dates = []

def validate_date(date_text):
    try:
        datetime.strptime(date_text, '%m/%d')
        return True
    except ValueError:
        return False

def add_task_dates(chronogram, start_date, ws, ws_project_schedule, ws_month, year, num_weeks, task_row_mapping, task_milestone_mapping, milestone_row_mapping, task_hours, row_offset=4):
    if not start_date:
        return None

    current_milestone = None
    global_start_date = datetime.strptime(f"{start_date}/{year}", "%m/%d/%Y")

    def get_next_available_date(date, used_dates):
        while date in used_dates:
            date += timedelta(days=7)
        return date

    used_start_dates = []
    used_end_dates = []

    milestone_week_hours = {}

    week_dates = get_week_dates(start_date, num_weeks, year)

    milestone_start_dates = {}
    milestone_end_dates = {}

    task_hours_index = 0

    for index, task_row in enumerate(chronogram):
        if set(task_row) == {''}:
            continue

        milestone_name = task_milestone_mapping[index]

        task_hour = task_hours[task_hours_index]
        task_hours_index += 1

        if current_milestone != milestone_name:
            current_milestone = milestone_name

            milestone_week_hours[current_milestone] = [40.0] * num_weeks

            if current_milestone and used_end_dates:
                last_end_date = max(used_end_dates)
                days_until_next_monday = (7 - last_end_date.weekday()) % 7 or 7
                global_start_date = last_end_date + timedelta(days=1)

        x_indices = [i for i, x in enumerate(task_row) if x == 'X']
        if x_indices and len(week_dates) > x_indices[0]:
            start_week_range, start_year = week_dates[x_indices[0]]
            end_week_range, end_year = week_dates[x_indices[-1]]

            start_week_range = start_week_range.split(' - ')[0]
            end_week_range = end_week_range.split(' - ')[1]

            if 'Dec' in start_week_range and 'Jan' in end_week_range:
                end_year += 1

            task_start_date = datetime.strptime(f"{start_week_range}/{start_year}", "%d/%b/%Y")
            task_end_date = datetime.strptime(f"{end_week_range}/{end_year}", "%d/%b/%Y")

            if task_start_date.month == 12 and task_end_date.month == 1:
                task_end_date = datetime.strptime(f"{end_week_range}/{start_year + 1}", "%d/%b/%Y")

            original_task_start_date = task_start_date
            original_task_end_date = task_end_date

            for i in range(x_indices[0], len(milestone_week_hours[current_milestone])):
                if task_hour <= milestone_week_hours[current_milestone][i]:
                    milestone_week_hours[current_milestone][i] -= task_hour
                    task_start_date = get_next_available_date(task_start_date, used_start_dates)
                    task_end_date = task_start_date + timedelta(days=6)
                    used_start_dates.append(task_start_date)
                    used_end_dates.append(task_end_date)
                    break
                else:
                    task_hour -= milestone_week_hours[current_milestone][i]
                    milestone_week_hours[current_milestone][i] = 0

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            if not isinstance(ws.cell(row=task_row_mapping[index], column=4), MergedCell):
                start_date_cell = ws.cell(row=task_row_mapping[index], column=4, value=original_task_start_date.strftime("%d-%b-%Y"))
                start_date_cell.border = thin_border
            if not isinstance(ws.cell(row=task_row_mapping[index], column=5), MergedCell):
                end_date_cell = ws.cell(row=task_row_mapping[index], column=5, value=original_task_end_date.strftime("%d-%b-%Y"))
                end_date_cell.border = thin_border

            if not isinstance(ws_project_schedule.cell(row=task_row_mapping[index], column=4), MergedCell):
                start_date_cell_ps = ws_project_schedule.cell(row=task_row_mapping[index], column=4, value=original_task_start_date.strftime("%d-%b-%Y"))
                start_date_cell_ps.border = thin_border
            if not isinstance(ws_project_schedule.cell(row=task_row_mapping[index], column=5), MergedCell):
                end_date_cell_ps = ws_project_schedule.cell(row=task_row_mapping[index], column=5, value=original_task_end_date.strftime("%d-%b-%Y"))
                end_date_cell_ps.border = thin_border

            if milestone_name not in milestone_start_dates or original_task_start_date < milestone_start_dates[milestone_name]:
                milestone_start_dates[milestone_name] = original_task_start_date
            if milestone_name not in milestone_end_dates or original_task_end_date > milestone_end_dates[milestone_name]:
                milestone_end_dates[milestone_name] = original_task_end_date

            for i, (date_range, date_year) in enumerate(week_dates, start=6):
                start_week_str, end_week_str = date_range.split(' - ')
                start_week = datetime.strptime(f"{start_week_str}/{date_year}", "%d/%b/%Y")
                end_week = datetime.strptime(f"{end_week_str}/{date_year}", "%d/%b/%Y")

                if start_week > end_week:
                    end_week = datetime.strptime(f"{end_week_str}/{date_year + 1}", "%d/%b/%Y")

                if original_task_start_date <= end_week and original_task_end_date >= start_week:
                    task_cell = ws.cell(row=task_row_mapping[index], column=i)
                    task_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                    task_cell.border = thin_border

    for milestone_name, start_date in milestone_start_dates.items():
        end_date = milestone_end_dates[milestone_name]
        milestone_row = milestone_row_mapping[milestone_name]
        start_date_cell = ws.cell(row=milestone_row, column=4, value=start_date.strftime("%d-%b-%Y"))
        start_date_cell.font = Font(bold=True)
        start_date_cell.border = thin_border

        end_date_cell = ws.cell(row=milestone_row, column=5, value=end_date.strftime("%d-%b-%Y"))
        end_date_cell.font = Font(bold=True)
        end_date_cell.border = thin_border

        ws_month.cell(row=milestone_row, column=4, value=start_date.strftime("%d-%b-%Y")).border = thin_border
        ws_month.cell(row=milestone_row, column=5, value=end_date.strftime("%d-%b-%Y")).border = thin_border

        for i, (date_range, date_year) in enumerate(week_dates, start=6):
            start_week_str, end_week_str = date_range.split(' - ')
            start_week = datetime.strptime(f"{start_week_str}/{date_year}", "%d/%b/%Y")
            end_week = datetime.strptime(f"{end_week_str}/{date_year}", "%d/%b/%Y")

            if start_week > end_week:
                end_week = datetime.strptime(f"{end_week_str}/{date_year + 1}", "%d/%b/%Y")

            if start_date <= end_week and end_date >= start_week:
                milestone_cell = ws.cell(row=milestone_row, column=i)
                milestone_cell.fill = PatternFill(start_color="32a852", end_color="32a852", fill_type="solid")
                milestone_cell.border = thin_border

                milestone_month_cell = ws_month.cell(row=milestone_row, column=i)
                milestone_month_cell.fill = PatternFill(start_color="32a852", end_color="32a852", fill_type="solid")
                milestone_month_cell.border = thin_border

    return None

def calculate_total_weeks(chronogram):
    max_length = max(len(row) for row in chronogram if set(row) != {''})
    return max_length

def adjust_column_settings(ws, ws_month, start_col_index, num_weeks, date_col_width=20):  # Increased width for demonstration
    column_widths = {
        'B': 7,
        'C': 30,
        'D': 12,
        'E': 12,
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
        ws_month.column_dimensions[col].width = width  # Adjust both ws and ws_month

    for i in range(num_weeks):
        col_letter = get_column_letter(start_col_index + i)
        ws.column_dimensions[col_letter].width = date_col_width
        ws_month.column_dimensions[col_letter].width = date_col_width  # Adjust both ws and ws_month

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=2, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    for row in ws_month.iter_rows(min_row=4, max_row=ws_month.max_row, min_col=2, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

def add_status_conditional_formatting(ws, start_row, end_row, col_index):
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    white_font = Font(color="FFFFFF")
    black_font = Font(color="000000")

    ongoing_dxf = DifferentialStyle(fill=green_fill, font=black_font)
    at_risk_dxf = DifferentialStyle(fill=yellow_fill, font=black_font)
    delayed_dxf = DifferentialStyle(fill=red_fill, font=white_font)

    ongoing_rule = Rule(type="containsText", operator="containsText", text="Ongoing", dxf=ongoing_dxf)
    ongoing_rule.formula = ['NOT(ISERROR(SEARCH("Ongoing",' + get_column_letter(col_index) + ')))']

    at_risk_rule = Rule(type="containsText", operator="containsText", text="At Risk", dxf=at_risk_dxf)
    at_risk_rule.formula = ['NOT(ISERROR(SEARCH("At Risk",' + get_column_letter(col_index) + ')))']

    delayed_rule = Rule(type="containsText", operator="containsText", text="Delayed", dxf=delayed_dxf)
    delayed_rule.formula = ['NOT(ISERROR(SEARCH("Delayed",' + get_column_letter(col_index) + ')))']

    ws.conditional_formatting.add(f"{get_column_letter(col_index)}{start_row}:{get_column_letter(col_index)}{end_row}", ongoing_rule)
    ws.conditional_formatting.add(f"{get_column_letter(col_index)}{start_row}:{get_column_letter(col_index)}{end_row}", at_risk_rule)
    ws.conditional_formatting.add(f"{get_column_letter(col_index)}{start_row}:{get_column_letter(col_index)}{end_row}", delayed_rule)

def process_final_week_ranges():
    global all_week_ranges
    return all_week_ranges

current_milestone = None
last_milestone_end_date = None
milestone_start_date = None
all_week_ranges = []
milestone_count = 0
current_milestone_count = 1
last_activity = None

def get_week_dates(start_date, num_weeks, year, milestone_name=None, last_end_dates=None, is_last_task=False):
    if not start_date:
        return [(f"Week {i+1}", year) for i in range(num_weeks + 1)]

    global last_milestone_end_date, current_milestone, milestone_start_date, all_week_ranges, current_milestone_count, milestone_count

    week_dates = []
    start_dates = []

    if last_milestone_end_date is not None and milestone_name != current_milestone:
        new_start_date = last_milestone_end_date + timedelta(days=1)
        start_dates = [new_start_date]
        milestone_start_date = new_start_date
    elif milestone_name == current_milestone and milestone_start_date:
        start_dates = [milestone_start_date]
    elif last_end_dates is not None:
        start_dates = [last_end_date + timedelta(days=1) for last_end_date in last_end_dates]
    if not start_dates:
        start_dates = [datetime.strptime(f"{start_date}/{year}", "%m/%d/%Y")]
        milestone_start_date = start_dates[0]

    current_dates = start_dates

    for i in range(num_weeks):
        end_dates = [current_date + timedelta(days=6) for current_date in current_dates]
        current_week_ranges = [
            f"{current_date.strftime('%d/%b')} - {end_date.strftime('%d/%b')}" for current_date, end_date in zip(current_dates, end_dates)
        ]
        week_dates.extend([(week_range, current_date.year) for week_range, current_date, end_date in zip(current_week_ranges, current_dates, end_dates)])
        all_week_ranges.extend([(week_range, current_date.year) for week_range, current_date, end_date in zip(current_week_ranges, current_dates, end_dates)])
        current_dates = [end_date + timedelta(days=1) for end_date in end_dates]

    if milestone_name:
        last_milestone_end_date = end_dates[-1] if end_dates else None
        current_milestone = milestone_name

    if current_milestone_count == milestone_count and milestone_name == current_milestone:
        process_final_week_ranges()

    return week_dates

def update_milestone_status(ws_project_schedule, milestone_row_mapping, last_filled_activity_task_row, row_checked):
    present_date = datetime.now()

    # Update the status of each task to 'Delayed' if its end date is past the current date
    for row in range(5, last_filled_activity_task_row + 1):
        end_date_cell = ws_project_schedule.cell(row=row, column=5).value
        status_cell = ws_project_schedule.cell(row=row, column=6)
        
        if end_date_cell:
            end_date = datetime.strptime(end_date_cell, "%d-%b-%Y")
            if end_date < present_date:
                status_cell.value = 'Delayed'
            else:
                status_cell.value = 'Ongoing'

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Update the status of each milestone based on the status of its tasks
    for milestone_name, milestone_row in milestone_row_mapping.items():
        delayed_tasks = 0
        total_tasks = 0
        
        # Calculate the end row for the current milestone
        milestone_end_row = next((row for row in range(milestone_row + 1, last_filled_activity_task_row + 1)
                                if ws_project_schedule.cell(row=row, column=2).value and
                                   "Task" in ws_project_schedule.cell(row=row, column=2).value and
                                   ws_project_schedule.cell(row=row, column=2).value.split('.')[0] != str(list(milestone_row_mapping.keys()).index(milestone_name) + 1)), 
                                last_filled_activity_task_row + 1)

        # Count the number of delayed and total tasks for the current milestone
        for row in range(milestone_row + 1, milestone_end_row):
            if ws_project_schedule.cell(row=row, column=2).value:  # Check if it's a task row
                total_tasks += 1
                status_cell = ws_project_schedule.cell(row=row, column=6)
                if status_cell.value == 'Delayed':
                    delayed_tasks += 1

        #print("")
        #print(f"Milestone: {milestone_name} - Delayed Tasks: {delayed_tasks}, Total Tasks: {total_tasks}")
        #print("")

        # Determine the milestone status
        if delayed_tasks == total_tasks and delayed_tasks > 0:
            milestone_status = 'Delayed'
        elif delayed_tasks > 0:
            milestone_status = 'At Risk'
        else:
            milestone_status = 'Ongoing'

        # Update the milestone status cell
        milestone_status_cell = ws_project_schedule.cell(row=milestone_row, column=6, value=milestone_status)
        milestone_status_cell.border = thin_border  # Add border to milestone status cell

        # Apply conditional formatting
        if milestone_status_cell.value == 'Ongoing':
            milestone_status_cell.fill = PatternFill(start_color="32CD32", end_color="32CD32", fill_type="solid")
            milestone_status_cell.font = Font(color="000000", bold=True)  # Black text for Ongoing
        elif milestone_status_cell.value == 'At Risk':
            milestone_status_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            milestone_status_cell.font = Font(color="000000", bold=True)  # Black text for At Risk
        elif milestone_status_cell.value == 'Delayed':
            milestone_status_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            milestone_status_cell.font = Font(color="FFFFFF", bold=True)  # White text for Delayed


def chronogramToExcel(chronogram, year, start_week, activity_names, milestoneNames, task_hours, filename="chronogram.xlsx"):
    start_col_index = 6
    num_weeks = calculate_total_weeks(chronogram)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    df = pd.DataFrame(chronogram)

    # Check if file is open
    if is_file_open(filename):
        print(f"File {filename} is open. Attempting to save with a new name.")
        timestamp = time.strftime("%Y%m%d-%H%M%S")
        filename = f"chronogram_{timestamp}.xlsx"

    try:
        df.to_excel(filename, index=False, header=False)
    except PermissionError:
        print(f"Permission denied for file {filename}. Attempting to save with a new name.")
        timestamp = time.strftime("%Y%m%d-%H%M%S")
        backup_filename = f"chronogram_{timestamp}.xlsx"
        df.to_excel(backup_filename, index=False, header=False)
        print(f"Permission denied. The Gantt Chart Excel file has been saved as {backup_filename}.")

    for col in range(start_col_index - 1):
        df.insert(col, 'Empty{}'.format(col), [''] * df.shape[0])
    df.to_excel(filename, index=False, header=False)

    if not year:
        week_years = set([date_info[1] for date_info in week_dates])
        if len(week_years) == 1:
            year = week_years.pop()
        else:
            year = min(week_years)

    wb = Workbook()
    ws = wb.active
    ws.title = "Gantt Chart (weeks)"

    ws_month = wb.create_sheet(title="Gantt Chart (months)")
    ws_project_schedule = wb.create_sheet(title="Project Schedule")  # Added Project Schedule sheet
    ws_raci_table = wb.create_sheet(title="RACI Table")  # Added RACI Table sheet

    format_blank_cells(ws)
    format_blank_cells(ws_month)
    format_blank_cells(ws_project_schedule)  # Format the new sheet
    format_blank_cells(ws_raci_table)  # Format the new RACI Table sheet

    headers = [("Tasks", 2), ("Activity", 3), ("Start Date", 4), ("End Date", 5), ("Status", 6)]
    project_schedule_headers = [("Tasks", 2), ("Activity", 3), ("Start Date", 4), ("End Date", 5),("Status", 6)]

    raci_headers = [("Tasks", 2), ("Activity", 3), ("Start Date", 4), ("End Date", 5), 
                    ("Product Owner", 6), ("Business Analyst", 7), ("Financial Lead", 8), ("Design Director", 9), ("CRM Lead", 10),
                    ("Head of CRM", 11), ("Senior Stakeholders* ", 12), ("Senior Stakeholders** ", 13), ("AGENCY ", 14)]  # Modified header for RACI Table

    for header, col in headers:
        for sheet in [ws, ws_month]:  # Loop through only ws and ws_month sheets
            if header == "Status":
                continue  # Skip Status for these sheets
            sheet.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
            header_cell = sheet.cell(row=1, column=col, value=header)
            header_cell.alignment = Alignment(horizontal='center', vertical='bottom')
            header_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            header_cell.font = Font(color="FFFFFF", bold=True)
            header_cell.border = thin_border

            for row in range(1, 4):
                for col in range(col, col + 1):
                    sheet.cell(row=row, column=col).border = thin_border

    for header, col in project_schedule_headers:
        sheet = ws_project_schedule  # Apply to the Project Schedule sheet
        sheet.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
        header_cell = sheet.cell(row=1, column=col, value=header)
        header_cell.alignment = Alignment(horizontal='center', vertical='bottom')
        header_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        header_cell.font = Font(color="FFFFFF", bold=True)
        header_cell.border = thin_border

        for row in range(1, 4):
            for col in range(col, col + 1):
                sheet.cell(row=row, column=col).border = thin_border

    for header, col in raci_headers:  # Apply modified headers to RACI Table
        sheet = ws_raci_table  # Apply to the RACI Table sheet
        sheet.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
        header_cell = sheet.cell(row=1, column=col, value=header)
        header_cell.alignment = Alignment(horizontal='center', vertical='bottom')
        header_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        header_cell.font = Font(color="FFFFFF", bold=True)
        header_cell.border = thin_border

        for row in range(1, 4):
            for col in range(col, col + 1):
                sheet.cell(row=row, column=col).border = thin_border



        


    row_offset = 5

    milestone_index = 0
    activity_index = 0
    task_index = 1
    task_row_mapping = {}
    task_milestone_mapping = {}
    milestone_row_mapping = {}

    new_chronogram = []
    new_activity_names = []

    for index, row in enumerate(chronogram):
        if set(row) == {''}:
            milestone_index += 1
            task_index = 1
            continue

        if milestoneNames[milestone_index] not in milestone_row_mapping:
            milestone_row_mapping[milestoneNames[milestone_index]] = len(new_chronogram) + row_offset
            new_chronogram.append([''] * len(row))
            new_activity_names.append(milestoneNames[milestone_index])

        task_label = f"Task {milestone_index + 1}.{task_index}"
        task_row_mapping[len(new_chronogram)] = len(new_chronogram) + row_offset
        task_milestone_mapping[len(new_chronogram)] = milestoneNames[milestone_index]
        task_index += 1

        new_chronogram.append(row)
        if activity_index < len(activity_names):
            new_activity_names.append(activity_names[activity_index])
            activity_index += 1

    milestone_counter = 0
    task_number = 1
    last_filled_activity_task_row = 0  # Initialize the variable
    row_checked = 0

    for index, row in enumerate(new_chronogram):
        excel_row = row_offset + index

        if set(row) == {''}:
            for sheet in [ws, ws_month, ws_project_schedule, ws_raci_table]:  # Loop through all sheets
                sheet.cell(row=excel_row, column=2, value=f"Task {milestone_counter + 1}")
                sheet.merge_cells(start_row=excel_row, start_column=2, end_row=excel_row, end_column=2)
                task_cell = sheet.cell(row=excel_row, column=2)
                task_cell.alignment = Alignment(horizontal='center', vertical='center')
                task_cell.font = Font(color="000000", bold=True)
                task_cell.border = thin_border

                sheet.cell(row=excel_row, column=3, value=milestoneNames[milestone_counter])
                sheet.merge_cells(start_row=excel_row, start_column=3, end_row=excel_row, end_column=3)
                cell = sheet.cell(row=excel_row, column=3)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(color="000000", bold=True)
                cell.border = thin_border

            milestone_counter += 1
            task_number = 1
        else:
            if index in task_row_mapping:
                task_excel_row = task_row_mapping[index]
                task_number_label = f"{milestone_counter}.{task_number}"
                for sheet in [ws, ws_project_schedule, ws_raci_table]:  # Loop through all sheets
                    task_cell = sheet.cell(row=task_excel_row, column=2, value=task_number_label)
                    task_cell.border = thin_border
                    sheet.cell(row=task_excel_row, column=3, value=new_activity_names[index])
                    sheet.cell(row=task_excel_row, column=3).border = thin_border

                    # Update the last filled row variable
                    last_filled_activity_task_row = task_excel_row

                    for col_index, value in enumerate(row, start=start_col_index):
                        task_cell = sheet.cell(row=task_excel_row, column=col_index)
                        if value == 'X':
                            if sheet != ws_project_schedule and sheet != ws_month and sheet != ws_raci_table:  # Exclude Project Schedule and RACI Table
                                task_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                                task_cell.border = thin_border
                task_number += 1

                # Ensure row_checked is set only for the current milestone
                if task_milestone_mapping[index] == milestoneNames[milestone_counter - 1]:
                    row_checked = task_number - 1

    adjust_column_settings(ws, ws_month, start_col_index, num_weeks, date_col_width=20)
    adjust_column_settings(ws_project_schedule, ws_month, start_col_index, num_weeks, date_col_width=20)  # Adjust the new sheet
    adjust_column_settings(ws_raci_table, ws_month, start_col_index, num_weeks, date_col_width=20)  # Adjust the RACI Table sheet

    add_task_dates(new_chronogram, start_week, ws, ws_project_schedule, ws_month, year, num_weeks, task_row_mapping, task_milestone_mapping, milestone_row_mapping, task_hours)
    add_task_dates(new_chronogram, start_week, ws, ws_raci_table, ws_month, year, num_weeks, task_row_mapping, task_milestone_mapping, milestone_row_mapping, task_hours)

    if not start_week:
        week_labels = [f"Week {i+1}" for i in range(num_weeks)]
        month_labels = [f"Month {i//4 + 1}" for i in range(num_weeks)]
        week_dates = [(f"Week {i+1}", year) for i in range(num_weeks)]
    else:
        week_dates = sorted(set(all_week_ranges), key=lambda x: (x[1], datetime.strptime(x[0].split(' - ')[0], '%d/%b')))

    if not week_dates:
        week_dates = get_week_dates("01/01", num_weeks, year)

    current_year = week_dates[0][1]
    year_start_col = start_col_index

    for i, (date_range, year_of_week) in enumerate(week_dates, start=start_col_index):
        if current_year != year_of_week:
            for sheet in [ws, ws_month]:  # Loop through ws and ws_month only
                sheet.merge_cells(start_row=1, start_column=year_start_col, end_row=1, end_column=i - 1)
                primary_cell = sheet.cell(row=1, column=year_start_col)
                primary_cell.value = str(current_year)
                primary_cell.alignment = Alignment(horizontal='left', vertical='center')
                primary_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
                primary_cell.font = Font(color="FFFFFF", bold=True)
                primary_cell.border = thin_border

                for col in range(year_start_col, i):
                    sheet.cell(row=1, column=col).border = thin_border

            current_year = year_of_week
            year_start_col = i

    for sheet in [ws, ws_month]:  # Loop through ws and ws_month only
        sheet.merge_cells(start_row=1, start_column=year_start_col, end_row=1, end_column=len(week_dates) + start_col_index - 1)
        primary_cell = sheet.cell(row=1, column=year_start_col)
        primary_cell.value = str(current_year)
        primary_cell.alignment = Alignment(horizontal='left', vertical='center')
        primary_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        primary_cell.font = Font(color="FFFFFF", bold=True)
        primary_cell.border = thin_border

        for col in range(year_start_col, len(week_dates) + start_col_index):
            sheet.cell(row=1, column=col).border = thin_border

    row_offset = 2
    months = {}

    actual_weeks_with_tasks = len(df.columns) - start_col_index + 1

    for i, (date_range, year) in enumerate(week_dates, start=start_col_index):
        if start_week:
            start_date_str, _ = date_range.split(' - ')
            try:
                start_date = datetime.strptime(start_date_str, "%d/%b")
                month_name = start_date.strftime("%B")
            except ValueError:
                month_name = "Unknown Month"

            if month_name not in months:
                months[month_name] = {'start': i, 'end': i}
            else:
                months[month_name]['end'] = i

        else:
            week_index = i - start_col_index
            month_num = (week_index // 4) + 1
            month_name = f'Month {month_num}'

            if month_name not in months:
                months[month_name] = {'start': i, 'end': i}

            if i < actual_weeks_with_tasks:
                months[month_name]['end'] = i

        for sheet in [ws, ws_month]:  # Loop through ws and ws_month only
            week_cell = sheet.cell(row=row_offset + 1, column=i)
            week_cell.value = date_range
            week_cell.alignment = Alignment(horizontal='center')
            week_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            week_cell.font = Font(color="FFFFFF")
            week_cell.border = thin_border

    if not start_week:
        for month_name, month_range in months.items():
            month_end_week = month_range['start'] + 3
            if month_end_week >= actual_weeks_with_tasks + start_col_index - 1:
                month_end_week = actual_weeks_with_tasks + start_col_index - 1
            if month_end_week < month_range['start']:
                month_end_week = month_range['start']
            months[month_name]['end'] = month_end_week

    for month_name, month_range in months.items():
        if month_range['start'] <= month_range['end']:
            for sheet in [ws, ws_month]:  # Loop through ws and ws_month only
                sheet.merge_cells(start_row=row_offset, start_column=month_range['start'], end_row=row_offset, end_column=month_range['end'])
                month_cell = sheet.cell(row=row_offset, column=month_range['start'])
                month_cell.value = month_name
                month_cell.alignment = Alignment(horizontal='center')
                month_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
                month_cell.font = Font(color="FFFFFF")
                for col in range(month_range['start'], month_range['end'] + 1):
                    sheet.cell(row=row_offset, column=col).border = thin_border

    for milestone_name, milestone_row in milestone_row_mapping.items():
        milestone_start_date = ws.cell(row=milestone_row, column=4).value
        milestone_end_date = ws.cell(row=milestone_row, column=5).value

        for sheet in [ws_month, ws_project_schedule, ws_raci_table]:  # Loop through all sheets
            sheet.cell(row=milestone_row, column=4, value=milestone_start_date).font = Font(bold=True)
            sheet.cell(row=milestone_row, column=4).border = thin_border

            sheet.cell(row=milestone_row, column=5, value=milestone_end_date).font = Font(bold=True)
            sheet.cell(row=milestone_row, column=5).border = thin_border

        for col in range(start_col_index, start_col_index + num_weeks):
            if ws.cell(row=milestone_row, column=col).fill.start_color.index == "32a852":
                for sheet in [ws_month, ws_project_schedule, ws_raci_table]:  # Loop through all sheets
                    sheet.cell(row=milestone_row, column=col).fill = PatternFill(start_color="32a852", end_color="32a852", fill_type="solid")
                    sheet.cell(row=milestone_row, column=col).border = thin_border

    # Add the data validation to the "RACI Status" column only for filled rows in the RACI Table sheet
    raci_status_validation = DataValidation(type="list", formula1='"Responsible,Accountable,Consulted,Informed"', allow_blank=True)
    raci_status_validation.error = 'Invalid entry, please select from the list'
    raci_status_validation.errorTitle = 'Invalid Entry'

    raci_status_col_start_index = 6  # Starting column for RACI status
    raci_status_col_end_index = 14  # Ending column for RACI status

    for row in range(5, last_filled_activity_task_row + 1):  # Apply up to the last filled row

        cell = ws_raci_table.cell(row=row, column=raci_status_col_start_index)
        cell.value = "Accountable"  # Set default value to "Informed"
        cell.border = thin_border  # Add border to the cell
        raci_status_validation.add(cell)

        for col in range(raci_status_col_start_index +1, raci_status_col_end_index):
            cell = ws_raci_table.cell(row=row, column=col)
            cell.value = "Informed"  # Set default value to "Informed"
            cell.border = thin_border  # Add border to the cell
            raci_status_validation.add(cell)

        cell = ws_raci_table.cell(row=row, column=raci_status_col_start_index+2)
        cell.value = "Consulted"  # Set default value to "Informed"
        cell.border = thin_border  # Add border to the cell
        raci_status_validation.add(cell)

        cell = ws_raci_table.cell(row=row, column=raci_status_col_end_index-4)
        cell.value = "Consulted"  # Set default value to "Informed"
        cell.border = thin_border  # Add border to the cell
        raci_status_validation.add(cell)

        cell = ws_raci_table.cell(row=row, column=raci_status_col_end_index-2)
        cell.value = "Consulted"  # Set default value to "Informed"
        cell.border = thin_border  # Add border to the cell
        raci_status_validation.add(cell)

        cell = ws_raci_table.cell(row=row, column=raci_status_col_end_index)
        cell.value = "Responsible"  # Set default value to "Informed"
        cell.border = thin_border  # Add border to the cell
        raci_status_validation.add(cell)

    ws_raci_table.add_data_validation(raci_status_validation)

    # Add conditional formatting rules for the RACI status columns
    for col in range(raci_status_col_start_index, raci_status_col_end_index + 1):
        col_letter = get_column_letter(col)
        ws_raci_table.conditional_formatting.add(f'{col_letter}5:{col_letter}{last_filled_activity_task_row}',
            CellIsRule(operator='equal', formula=['"Responsible"'], fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"), font=Font(color="FFFFFF", bold=True)))  # Red fill with white text
        ws_raci_table.conditional_formatting.add(f'{col_letter}5:{col_letter}{last_filled_activity_task_row}',
            CellIsRule(operator='equal', formula=['"Accountable"'], fill=PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"), font=Font(color="000000", bold=True)))  # Yellow fill with black text
        ws_raci_table.conditional_formatting.add(f'{col_letter}5:{col_letter}{last_filled_activity_task_row}',
            CellIsRule(operator='equal', formula=['"Consulted"'], fill=PatternFill(start_color="800080", end_color="800080", fill_type="solid"), font=Font(color="FFFFFF", bold=True)))  # Purple fill with white text
        ws_raci_table.conditional_formatting.add(f'{col_letter}5:{col_letter}{last_filled_activity_task_row}',
            CellIsRule(operator='equal', formula=['"Informed"'], fill=PatternFill(start_color="008000", end_color="008000", fill_type="solid"), font=Font(color="FFFFFF", bold=True)))  # Green fill with white text



    # Apply the data validation to the "Status" column only for filled rows
    status_validation = DataValidation(type="list", formula1='"Ongoing,At Risk,Delayed"', allow_blank=True)
    status_validation.error = 'Invalid entry, please select from the list'
    status_validation.errorTitle = 'Invalid Entry'

    status_col_index = 6  # Assuming the "Status" column is at index 6

    present_date = datetime.now()

    for row in range(5, last_filled_activity_task_row + 1):
        end_date_cell = ws_project_schedule.cell(row=row, column=5).value
        status_cell = ws_project_schedule.cell(row=row, column=6)
        
        if end_date_cell:
            end_date = datetime.strptime(end_date_cell, "%d-%b-%Y")
            if end_date < present_date:
                status_cell.value = 'Delayed'

    # Apply the data validation to the "Status" column only for filled rows
    status_validation = DataValidation(type="list", formula1='"Ongoing,At Risk,Delayed"', allow_blank=True)
    status_validation.error = 'Invalid entry, please select from the list'
    status_validation.errorTitle = 'Invalid Entry'

    # Add the data validation to the "RACI Status" column only for filled rows in the RACI Table sheet
    raci_status_validation = DataValidation(type="list", formula1='"Responsible,Accountable,Consulted,Informed"', allow_blank=True)
    raci_status_validation.error = 'Invalid entry, please select from the list'
    raci_status_validation.errorTitle = 'Invalid Entry'

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    status_col_index = 6  # Assuming the "Status" column is at index 6
    for row in range(5, last_filled_activity_task_row + 1):  # Apply up to the last filled row
        cell = ws_project_schedule.cell(row=row, column=status_col_index)
        if not cell.value:
            cell.value = 'Ongoing'
        cell.border = thin_border  # Add border to the cell
        status_validation.add(cell)

    ws_project_schedule.add_data_validation(status_validation)
    #ws_raci_table.add_data_validation(status_validation)  # Apply validation to RACI Table as well


    # Update milestone status
    update_milestone_status(ws_project_schedule, milestone_row_mapping, last_filled_activity_task_row, row_checked)
    #update_milestone_status(ws_raci_table, milestone_row_mapping, last_filled_activity_task_row, row_checked)  # Update status in RACI Table as well

    # Add conditional formatting rules for the status column
    for sheet in [ws_project_schedule]:  # Apply to both sheets
        sheet.conditional_formatting.add(f'F5:F{last_filled_activity_task_row}',
            CellIsRule(operator='equal', formula=['"Ongoing"'], fill=PatternFill(start_color="32CD32", end_color="32CD32", fill_type="solid"), font=Font(color="000000", bold=True)))  # Black text
        sheet.conditional_formatting.add(f'F5:F{last_filled_activity_task_row}',
            CellIsRule(operator='equal', formula=['"At Risk"'], fill=PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"), font=Font(color="000000", bold=True)))  # Black text
        sheet.conditional_formatting.add(f'F5:F{last_filled_activity_task_row}',
            CellIsRule(operator='equal', formula=['"Delayed"'], fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"), font=Font(color="FFFFFF", bold=True)))  # White text

    # Get the current working directory
    current_directory = os.getcwd()

    try:
        df.to_excel(filename, index=False, header=False)
        wb.save(filename)
        df.to_csv("chronogram.csv", index=False)
        print("The excel file has been generated in the directory:", current_directory, "\n")
        print(f"The Gantt Chart Excel file has been successfully generated as {filename}.")
    except PermissionError:
        print(f"Permission denied for file {filename}. Attempting to save with a new name.")
        timestamp = time.strftime("%Y%m%d-%H%M%S")
        backup_filename = f"chronogram_{timestamp}.xlsx"
        wb.save(backup_filename)
        df.to_excel(backup_filename, index=False, header=False)
        df.to_csv(f"chronogram_{timestamp}.csv", index=False)
        print("The excel file has been generated in the directory:", current_directory, "\n")
        print(f"Permission denied. The Gantt Chart Excel file has been saved as {backup_filename}.")


yearInput = input("Add the year for the Gantt Chart (leave empty if using current year):\nInput: ").strip()
year = int(yearInput) if yearInput else datetime.now().year

print()

start_week = input("Add the starting week (MM/DD) (leave empty if not):\nInput: ").strip()
while start_week and not validate_date(start_week):
    start_week = input("The format is incorrect. Please use MM/DD format or leave empty:\nInput: ").strip()

print() 

milestoneNames = []
milestonesInput = input("Enter the list of milestones (as comma-separated values), or leave empty:\nInput: ")
if milestonesInput:
    milestoneNames = [milestone.strip() for milestone in milestonesInput.split(',')]
    milestone_count = len(milestoneNames)
else:
    milestone_count = 0

print()  

activityNames = []
milestones_tasks = []
task_hours = []

for index, milestone in enumerate(milestoneNames):
    print()
    print(f"Adding tasks for Milestone: {milestone}")

    tasksInput = input(f"Enter the list of tasks for {milestone} (as comma-separated values):\nInput: ")
    while not tasksInput:
        tasksInput = input(f"Add at least one task for {milestone} (as comma-separated values):\nInput: ")

    tasks = [task.strip() for task in tasksInput.split(',')]
    while not all(tasks):
        print("Task names can't be empty. Please enter valid task names.")
        tasksInput = input(f"Enter the list of tasks for {milestone} (as comma-separated values):\nInput: ")
        tasks = [task.strip() for task in tasksInput.split(',')]

    print()  

    #print()  

    taskHoursInput = input(f"Enter the hours for tasks under {milestone} (as comma-separated values):\nInput: ")
    while not taskHoursInput or not all(re.match(r'^\d+(\.\d+)?$', x.strip()) for x in re.split(r'[,\s]+', taskHoursInput) if x.strip()):
        print("Input format is incorrect. Please enter only numbers (integers or floats) separated by commas or spaces.")
        taskHoursInput = input(f"Add at least one task hour for {milestone} (as comma-separated values):\nInput: ")

    hours = [float(x.strip()) for x in re.split(r'[,\s]+', taskHoursInput) if x.strip()]

    milestoneActivityNames = [f"{task}" for task in tasks]
    activityNames.extend(milestoneActivityNames)
    milestones_tasks.append((milestone, hours))

    task_hours.extend(hours)

    if index == len(milestoneNames) - 1:
        last_activity = milestoneActivityNames[-1] if milestoneActivityNames else None

print("\n") 

chronogram = allocateTasksToWeeks(milestones_tasks)
chronogramToExcel(chronogram, year, start_week if start_week.strip() else "", activityNames, milestoneNames, task_hours, "chronogram.xlsx")
